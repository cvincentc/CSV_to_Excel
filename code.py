import xlrd
import csv
import sys
import getopt
import pandas as pd
from datetime import datetime
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl import load_workbook
import xlsxwriter
import openpyxl


#Styling for excel file
def format_fn(file):
    #Open file
    wb = load_workbook(file)

    #Set font
    my_font = Font(name = 'Calibri',
                   size=15,
                   bold=True,
                   italic=False,
                   vertAlign=None,
                   underline='none',
                   strike=False,
                   color='000000')

    #Get sheet
    sheet = wb['Sheet1']

    #Set border style
    b_style = Side(border_style="thin",color="000000")
    border = Border(top=b_style,left=b_style,right=b_style,bottom=b_style)

    #Set column width
    sheet.column_dimensions['A'].width=5
    sheet.column_dimensions['B'].width=5.5
    sheet.column_dimensions['C'].width=4
    sheet.column_dimensions['D'].width=4
    sheet.column_dimensions['E'].width=65
    sheet.column_dimensions['F'].width=21
    sheet.column_dimensions['G'].width=9
    sheet.column_dimensions['H'].width=4

    #Alternating shading for more clear representation
    my_shade = openpyxl.styles.colors.Color(rgb='E0E0E0')
    my_fill = openpyxl.styles.fills.PatternFill(patternType="solid", fgColor=my_shade)

    #Set styling for every cell
    for row_cells in sheet.iter_rows():
        for cell in row_cells:
            cell.font = my_font
            cell.border=border
            
            if(cell.row%2==0):
                cell.fill = my_fill
                
            if cell.column==5:
                
                #Shortening address
                index = str(cell.value).find('British')
                if(index<1): continue
                cell.value=(str(cell.value))[:index-2]
                
    #Save file
    wb.save(file)


#main function
def main(arg):
    now = datetime.now()
    
    
    #Read Onfleet excel data 
    df = pd.read_csv(arg,usecols=[1,7,10])

    #Get driver name
    driver = pd.read_csv(arg,usecols=[6]).iloc[1]['workerName']

    #Header
    s = now.strftime("%Y-%m-%d")
    s+= ' ' + driver + ' Delivering List'
    df.columns=['',s ,'Phone#']

    #Insert header
    df.insert(1,'L','')
    df.insert(2,'G','')
    df.insert(3,'F','')
    df.insert(int(df.iloc[1].count()),'Arrived','')
    df.insert(int(df.iloc[1].count()),'PU','')

    new_row1 = pd.Series(["Total:"],index=[''])
    new_row2 = pd.Series(['Note:'],index=[''])

    #Insert additional rows
    df=df.append(new_row1,ignore_index=True)
    df=df.append(new_row2,ignore_index=True)

    pd.set_option("display.max_rows", None, "display.max_columns", None)
    
    current_time = now.strftime("%d%H%M%S")
    filename = current_time+'.xlsx'

    #Save to new file
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')    
    df.to_excel(writer,sheet_name='Sheet1',index=False)
    writer.save()

    #Add styling
    format_fn(filename)


if __name__ == "__main__":
    main(sys.argv[1])
    
